# OGOCHUKWU JANE OKAFOR

# VOTING RULES

import numpy as np
import openpyxl

def generatePreferences(values):
    """
    Generate preferences function will generate a preference profile showing the agents and their favourable alternative/preference down to the second- most, second to the last and the least favourable preferences.
    However, the end result will be  a dictionary which will show agents as keys and positions of each alternatives (items/values).
    Here, only the indexes of the alternatives of each alternative will be return, but they will be sorted based on the highetst to the minimum values.
    Finally, since indexes start from 0, 1 is added so that the first alternative as well as the first agents will count from 1, while the second is 2 and so on.
    
    Parameters:
        values: a worksheet- openpyxl.load_workbook('voting.xlsx') 

    Returns:
        dict:   the preferences profile; that is the dictionary created to store the votings by the six agents, starting from an alternative with the highest value down to the alternative with the lowest value.
    """

    # n = sets of n agents
    # m = sets of m alternatives
    # agents = values.rows
    # alternatives = values.columns
    # agents are rows = 6
    # alternatives are columns = 4 per row
    # make a list, the hightest comes first followed by the 2nd and the rest
   
    agent = 0 # agent index is zero at first
    # creating an empty dictionary for preferences profile
    my_Prefs_dict = {} # preference profile
    
    for row in values: # in order to access the openpyxl worksheet, we need the six active rows
        agent +=1 # since the values in the first row is the 0 index while the latter colums are index 1, 2 and 3, 1 is added to each row of the agents.
        agent_Pref = [] # a list to put the four preference options or alterantives made by the each agents respectively.
        agent_valuations = np.array([cell.value for cell in row]) # using numpy here in order to store the list of the values of the agent preferences
        sorted_valuations = sorted(agent_valuations, reverse=True) # sorting the agent list in order to reaffarnge their positions based on values with the highest points.
  
        for valuation in sorted_valuations: # for values in the sorted valuations as seen above
            alternatives = np.where(agent_valuations == valuation) # alternatives here are four respective agent valuations in each row.
            for pos in alternatives[0][::-1]: # for positions in alternatives
                if pos + 1 in agent_Pref: # if position is already existing in the agent prefernce list
                    pass # this will be ignored since the position is already existing

                else:
                    agent_Pref.append(pos + 1) # append th agent preference list by adding the alternative point inside it.
        
        my_Prefs_dict[agent] = agent_Pref # the parameters of the preference profile dictionary is the list of the agent preferences
   
    return my_Prefs_dict # the preferences profile dictionary is returned in the execution of this function.
    
def dictatorship(preferenceProfile, agent):
    """
    Dictatorship function will determine the alternative that wins based on the alternative that is ranked first on the list.

    Parameters:
        preferenceProfile:  the dictionary which holds the four respective preferences of the each agent.
        agent:              comprises the six rows which have a preference order of four alternatives as columns.

    Returns:
        int -               the winner is returned based as dictatiorship rule function
    """
    if not isinstance(agent, int): # if an agent is not an integer, e.g a float or a string
        raise TypeError("The agent is not an integer! ") # a type error "the agent is not an integer" will be returned
    
    elif agent not in preferenceProfile.keys(): # else, if agent is not among the preference profile keys
        raise ValueError("The agent cannot be seen in the preference profile") # a value error will be returned since the agent cannot be found on the preference profile.
          
    else:
        return preferenceProfile[agent][0]  #will return the agent in to the preference profile if the agent can be seen in it.


def scoringRule(preferences, scoreVector, tieBreak):
    """
    Scoring Rule function will determine a winner based on an alternative with the highest total score and selects a winner if there are one or more alternatives using tie breaking.

    Parameters:
        preferences:    a dictionary that holds the preferences profile of all agents and their alternatives
        scoreVector:    a list of floting numbers
        tieBreak:       string ("max", "min" and "agent i")
            int-        agent that declares a winner

    Returns:
        int-            returns the alternative who has the highest total score. 
                        If we have more than one alternative with the same total score, tie-breaking comes in here to select the winner.
    """
    #Making sure the length of the scoreVector is correct/right
    if len(scoreVector) != len(preferences[1]): # if the length of score vector is not even to the length of the preferences profile. Note, one agent must exist.
        print("Incorrect Input") # Incorrect input will be printed out if these two lengths above are not equal
        return False # will return false, not true.

    else:
        alt_scores = {} # a dictionary holding alternative scores
        sorted_points = sorted(scoreVector, reverse=True) # a sorting the score points for the score vector for the socing rule

        for agent in preferences.keys(): # for every agent that are keys of the preferences profile
            score_index = 0 # score index reads from the 0 index
            for alternative in preferences[agent]: # for every agent who ar agents of the preference's agent
                if alternative not in alt_scores.keys(): # if an alternative is not in the dictionary holidng the existing in the alternativ scores 
                    alt_scores[alternative] = sorted_points[score_index] # alternative in the alt_scores dictionary is equal to the sorted index of the sorted points as seen above.
                
                else:
                    alt_scores[alternative] += sorted_points[score_index] # if an alternative is in the dictionary holidng the existing in the alternativ scores 
                score_index +=1  # increase  the score index to one
        altern_max_pt = [] # a list to hold all alternatives maximum points
        max_pt = max(alt_scores.values()) # assigning max_pt to defined the maximum value of all the values in the alt_scores dictionary

        for alternative in alt_scores.keys(): # for every altetantive as keys in the alt_scores dictionary
            if alt_scores[alternative] == max_pt: # max_pt is true to be equal to the alternative in the alt_scores dictionary
                altern_max_pt.append(alternative) # append the altern_max_pt list by adding the alternative inside it.
        
        return tieBreaking(altern_max_pt, tieBreak, preferences) # returns the tiebreaking rule in order to determine the winner if there is a tie.


def tieBreaking(alternList, tieBreak, preferences):
    """
    Tie - Breaking function will determine a winner based on the three rules
    If an alternative has the highest or lowest scores, or an agent is an interger, this fucntion will execute who the winner will be.

    Parameters:
        alternList:     a list of alternatives 
        tieBreak:       string ("max", "min" and "agent i")
             int-       agent that declares a winner
        prreferences:   a dictionary of the preferences profile

    Returns:
        int-            a winner is returned based on the tie breaking rule as seen above.
    """
    if tieBreak == "max": # if tie break is the alternative with the maximum index
            return max(alternList) # returns the first alternative

    elif tieBreak == "min": # if tie break is the alternative with the minimum index
            return min(alternList) # returns the alternative who has the the minimum index

    # Find the alternative with the highest score
    elif tieBreak in preferences.keys(): # if there is tie-breaking in the preferences' keys
        highest_index = len(preferences[tieBreak]) - 1 # subtract one from the length of the preferences profile
        for alternative in alternList: # for every alternatives in the alternative lsit
            if preferences[tieBreak].index(alternative) <= highest_index: # if the tiebreak in preferences's alternative index  is less than the length of the preferene profile
                winner = alternative  # the winner is the alternative
                highest_index = preferences[tieBreak].index(alternative) # the length of the prefernces profile equals the tiebreak in preferences's alternative index

        return winner # the winner is returned 

    else: 
        raise ValueError("Invalid Tie-Breaking Rule") # a value error is raised if this occurs


def plurality(preferences, tieBreak):
    """
    Plurality funcion will determine a winner by finding the the alternatives who have first positions in the preferences profile according to the plurality function.
    
    Parameters:
        preferences:    a dictionary of the preference profile of all alternatives by the six agents.
        tieBreak:       string ("max", "min" and "agent i")
            int -       agent in the the preference dictionary that declares a winner.
    
    Returns:
        int -           the winner using the plurality rule
    """
    alt_freq = {} # a dictionary of alternative frequencies
    first_positions = [] # a list of alternatives with first positions
    for agent in preferences.keys(): # for every agent existing in preferences profile
        first_pts = preferences[agent][0] # at first, first position is zero
        first_positions.append(first_pts) # append the first posiion list by adding the first_pts to it

    for alternative in first_positions: # for every alternative in the list of first positions
        if alternative in alt_freq.keys(): # if teh alternative exists in the alternative frequencies dictionary
            alt_freq[alternative] += 1 # the alterantives in the alt_freq dictionary is increased/equal to 1
        else:
            alt_freq[alternative] = 1 # the alterantives in the alt_freq dictionary is equal to 1
    
    max_counts = max(alt_freq.values()) # finding the alternative with maximum number of occurences

    alt_win =[]  # placing alternatives ranking 1st positions in a list
    for alternative in alt_freq: # for all alternatives in the alt_freq dictionary
        if alt_freq[alternative] == max_counts: # alternatives in alt_freq dict is true to be equal to the maximum number of occurences
            alt_win.append(alternative) # append the alternatives ranking at first positions inside the alt_win list

    return tieBreaking(alt_win, tieBreak, preferences) # if a tie exist, the tie breaking rule play a decision making here

def veto(preferences, tieBreak):
    """
    Veto function will determine a winner based on the veto function rule .
    Every agent assigns 0 points to the alternative that ranks in the last place of their preferences and 1 point to every other alternative.

    Parameters:
        preferences:    a dictionary of the preferences profile
        tieBreak:       string ("max", "min" and "agent i")
            int-        agent that declares a winner

    Returns:
        int -           returns a winner based on the veto rule function keeping in mind the socirng rule as regards tie breaking
    """
    #Find length of alternative list.
    m = len(list(preferences.items())[0][1]) #assigning m to be the length of the alternative list
    #create a list called score vector. make the last value zero, the rest one
    #if i = m -1, append 0 into the score vector, else append 1
    scoreVector = [] # a floating list of the score vector
    for i in range(m): # for every alternative in sets of m alternatives
        if i == m - 1: #m - 1  is the last index
            scoreVector.append(0) # append the score vector list by adding zero
        else:
            scoreVector.append(1) # append the score vector list by adding one

    return scoringRule(preferences, scoreVector, tieBreak) # returns a winner based on the score rule function

def borda(preferences, tieBreak): 
    """
    Borda rule function will determine a winner based on agents' preferences and considers the tie breaking rules.
    Each agent assigns 0 score of to its least-preferred alternative, the alternatives ranking at the bottom of the preferenes profile.
    Each agent assigns 1of  to the second least-preferred alternative, ... , and a score of  to their favourite alternative.

    Parameters:
        preferences:    a dictionary of the preferences profile
        tieBreak:       string ("max", "min" and "agent i")
             int-       agent that declares a winner

    Returns:
        int- returns a winner based on the borda function rule and takes note of the scoring rule as determined by tie breaking.
    """
    m = len(list(preferences.items())[0][1]) # lets assign m to be the length of preferences that are items in the list
    scoreVector = [i for i in range(m)] # assigning score vector to define every item seen in the range m
    
    return scoringRule(preferences, scoreVector, tieBreak) # returns the winner based on score ruling rule and keeps in mind the tie-breaking rules

def harmonic(preferences, tieBreak):
    """
    Harmonic rule function will determine a winner based on the alternative with the highest score.
    1/m is assigned to least preferred alternative, 1/m - 1, to the second least preferred alternative while 1 to the most preferred alternative.

    Parameters:
        preferences:    a dictionary of the preferences profile
        tieBreak:       string ("max", "min" and "agent i")
             int-       agent that declares a winner

    Returns:
        int- returns a winner based on the borda function rule and takes note of the scoring rule as determined by tie breaking.
    """
    m = len(list(preferences.items())[0][1]) # assigning m to be the length of preferences that are items in the list
    scoreVector = [] # creating an empty of score vector to keep records of scores

    for i in range(m): # for every item seen in the range m
        scoreVector.append(1/(i+1)) # appends the score vector lsit by adding 1/(i + 1); where i is 4 (alternatives 1-4)

    return scoringRule(preferences, scoreVector, tieBreak) # returns the winner based on score ruling rule while considering the tie-breaking rules


def STV(preferences, tieBreak):
    """
    Single Transferable Vote function elects a single winner from the preference ordering of alternatibves. 
    It works by allowing agents to rank the candidates in order of preference.
    Removes alternatives who has the least counts of first/favourite positions from the preference profile.
    If an alternative appeared once or not at all at the first place of the preference orderings, it will be the first one to be kicked out.

    Parameters:
        preferences:    a dictionary of preferences profile
        tieBreak:       string ("max", "min" and "agent i")
            int-        agent that declares a winner

    Returns:
        int-            a winner based on STV rule function and keeps in mind the tie breaking rule if there are multiple possible winners
    """
    # get a duplicated copy of the preferences profile
    dummy_pref = {} # creating an empty dictionary to be a duplicate/copy of the preferences profile since it is not advisable to modify the original dictionary.

    for key in preferences.keys(): # for every key existing as keys in the preferences profile dictionary
        dummy_pref[key] = list(preferences[key]) # coping the list of keys of the preferences dictionary to be the keys of the newly created duplicate/copy dictionary
    
    max_freq = 1 # maximum frequency equals one
    min_freq = 2  # minimum frequency equals two

    while (max_freq != min_freq): # while loop when maximum frequency is true to be equal to minimum frequency
        #keep eliminating

    # get dictionary
        alter_freq = {} # creating an empty dictionary to store alterative frequencies
        for alternative  in dummy_pref[1]: # for every alternative  existing in the duplicated dictionary
            alter_freq[alternative] = 0 # alternative in the alternative frequencies dictionary at first is 0
        
        for agent in dummy_pref.keys(): # for every agent existing in the duplicated dictionary
            # Lets agent of the duplicated dictionary for alternative frequennceis to be += 1
            alter_freq[ dummy_pref[agent][0] ] += 1 # since index starts from 0, add 1 to the 1st index 

        max_freq = max(alter_freq.values()) # maximum values of the alternative frequencies dictionary
        min_freq = min(alter_freq.values()) # minimum values of the alternative frequencies dictionary
        
        # creating an empty list for alternatives maximum points
        alter_mx_pt =[] #Holds the maximum points of each alternative
        if (max_freq == min_freq): # if the maximum frequnecy is true to be qual to the minimum frequency
            for alt in alter_freq.keys(): # for every alternative that are keys of the alter_freq dictionary
                alter_mx_pt.append(alt) # append the alternative into the list holding the maximum points of each alternative
                break # a condition to for a stop when this condition is met, otherwise, may crash the system  
        
        else:
            # creating an empty elimination list
            elim =[] # this list will keep records of the possible winners
            for alternative in alter_freq.keys(): # for every alternative  existing in the alter_freq dictionary
                if alter_freq[alternative] == min_freq: # if the alternatives seen in the alter_freq dictionary is true to be equal to the minimum frequenecy
                    elim.append(alternative) # append the elimination list by adding the alternative into it

            for alternative in elim: # for every alternative existing in the elimination list
                for agent in dummy_pref.keys(): # for every agent who is a key of the duplicated preferences profile
                   dummy_pref[agent].remove(alternative) # remove/elimiante this alternative from each agent existing in the duplicated preferences profile
            
            for agent in dummy_pref.keys(): # for every agent who is a key of the duplicated preferences profile
                if dummy_pref[agent][0] in alter_mx_pt: # if an agent of the duplicated preferences profile who is at index 0 is in the list holding the maximum points of each alternative
                    pass # skipped or pass, do not append

                else: 
                    # otherwise
                    alter_mx_pt.append(dummy_pref[agent][0]) # append the list holding the maximum points of each alternative by adding the dummy_pref's agent [0] to it
            
            #creating an empty dictionary to hold alternatives frequencies
            alter_freq = {}
            for alternative  in dummy_pref[1]: # for every alternative existing in the newly created alternative frequencies dictionary
                alter_freq[alternative] = 0 # at first, the alternatives in the alternative frequency dictionary is 0
            
            for agent in dummy_pref.keys(): # for every agent who is a key of the duplicated preferences profile
                alter_freq[ dummy_pref[agent][0] ] += 1 # the dict of the alternative frequencies holding the agents of the duplicated preferences profile is plus or equals to 1

            max_freq = max(alter_freq.values()) # maximum values of the alternative frequencies dictionary
            min_freq = min(alter_freq.values()) # minimum values of the alternative frequencies dictionary
  
    return tieBreaking (alter_mx_pt,tieBreak,preferences) # returns the winner based on the rules of tiebreaking
    

def rangeVoting(values, tieBreak):
    """
    Range voting determines a winner based on the range voting rule function.
    All alternatives ranging from first to the least in the preference orderings of different agents are summed up as regards to their positions.

    Parameters: 
        values:     a worksheet- openpyxl.load_workbook('voting.xlsx') 
        tieBreak:   string ("max", "min" and "agent i")
            int-    agent that declares a winner

    Returns:
        int -       returns a winner based on range voting function and keeps in mind tie breaking scoring rule if there are multiple possible winners.
    """
    # values = openpyxl.load_workbook('voting.xlsx') # will access the voting.xlsx spreadsheet
    
    alt_tp = {} # a dictionary holding alternatives total points
   
    for row in values: # for every row in the worksheet
        alt = 1 # there must exist atleast one alternative
        for cell in row: # for each alternative or column existing in each row or agent
            if alt not in alt_tp.keys(): # if an alternative is not in the key of the total points of the alternatives
                alt_tp[alt] = cell.value # alternatives in the total points of the alternatives equals each cells existing in the worksheet

            else:
                alt_tp[alt] += cell.value # alternatives in the total points of the alternatives plus or equal to each cells existing in the worksheet
            alt += 1 # alternative plus or equals to 1
    
    alter_mx_pt = [] # a list of the maximum points oe every single alternative
    mx_pt = max(alt_tp.values()) # assigns mx_pt to define the maximum sum of valuations as shown in the alt_tp dictionary

    for alt in alt_tp.keys(): # for every alternative that are keys in the dictionary (alt_tp) holding alternatives total points
   
        if alt_tp[alt] == mx_pt: # alternatives in the dict alt_tp is true to be equal to the maximum sum of valuations
            alter_mx_pt.append(alt) # append the alter_mx_pt list by adding the alternative to it
    
    preferences = generatePreferences(values) # the preferences here are the preferenced derived from the generatePreferences function

    return tieBreaking(alter_mx_pt, tieBreak, preferences) # returns the winner based on the rules of tiebreaking
  

if __name__ == "__main__":
    vote = openpyxl.load_workbook('voting.xlsx') #Will open the voting.xlsx spreadsheet
    votings = vote.active #will access the active workbook/spreadsheet
    #for row in votings.rows
    # for col in votings.columns
    #totalNumOfVotes = 0 #Finding out the total number of votes in the sheet
    #preferences = agent_Pref
    sample = generatePreferences(votings)
    print(sample)
    scoringRule(sample,[3,2.4,3.5,1.0],"max")
    w = STV(sample,"max")
    pref ={
    1:[1,2,3,4,5,6],
    2:[4,5,1,3,2,6],
    3:[6,1,2,3,4,5],
    4:[5,3,2,1,4,6],
    5:[1,3,2,4,5,6],
    6:[3,2,4,6,5,1],
    7:[3,1,4,5,6,2],
    8:[1,3,2,5,4,6],
    9:[1,6,4,5,3,2],
    10:[6,3,2,4,5,1]}

    k = STV(pref,"max")
    print(k)
